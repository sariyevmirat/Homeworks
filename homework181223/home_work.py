from openpyxl import Workbook
import openpyxl as op
def sort (x):
    filename = x
    wb = op.load_workbook(filename, data_only=True)
    sheet = wb.active

    all_values = []

    for row in sheet.iter_rows(values_only=True):
        for value in row:
            if value is not None:
                all_values.append(value)

    sortirovka = sorted(all_values, reverse=True)
    return sortirovka

wb = Workbook()
ws = wb.create_sheet('книга',0)
fn = '1.xlsx'

a = (1111, 2222, 3333)
ws.append(a)
wb.save(filename=fn)

sorted_data = sort(fn)
wb_fn = op.load_workbook(fn)
ws_fn = wb_fn.active

row_index = 1
for value in sorted_data:
    ws_fn.cell(row=2, column=row_index, value=value)
    row_index += 1

wb_fn.save(filename=fn)

wb1 = Workbook()
ws1 = wb1.create_sheet('книга',0)
fn1 = '2.xlsx'

b = (1111, 2233322, 33222233)
ws1.append(b)
wb1.save(filename=fn1)

sorted_data = sort(fn1)
wb_fn = op.load_workbook(fn1)
ws_fn = wb_fn.active

row_index = 1
for value in sorted_data:
    ws_fn.cell(row=2, column=row_index, value=value)
    row_index += 1

wb_fn.save(filename=fn1)

wb2 = Workbook()
ws2 = wb2.create_sheet('книга',0)
fn2 = '3.xlsx'

c = (555, 22212312312, 33434343433)
ws2.append(c)
wb2.save(filename=fn2)

sorted_data = sort(fn2)
wb_fn = op.load_workbook(fn2)
ws_fn = wb_fn.active

row_index = 1
for value in sorted_data:
    ws_fn.cell(row=2, column=row_index, value=value)
    row_index += 1

wb_fn.save(filename=fn2)

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side

wb_combined = Workbook()
ws_combined = wb_combined.active

font_style = Font(name='Arial', size=15, bold=True)

border_style = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))

files = ['1.xlsx', '2.xlsx', '3.xlsx']
for i in files:
    wb = load_workbook(i, data_only=True)
    sheet = wb.active

    for row in sheet.iter_rows(values_only=True):
        ws_combined.append(row)

for row in ws_combined.iter_rows():
    for cell in row:
        cell.font = font_style
        cell.border = border_style


wb_combined.save('файл.xlsx')
wb2.close()
wb1.close()
wb.close()