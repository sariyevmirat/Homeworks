import openpyxl

workbook = openpyxl.Workbook()
sheet = workbook.active

for i in range(1, 11):
    sheet.cell(row=1, column=i, value=i)

for i, letter in enumerate('ABCDEFGHIJ', start=1):
    sheet.cell(row=2, column=i, value=letter)

workbook.save('original_data.xlsx')

loaded_workbook = openpyxl.load_workbook('original_data.xlsx')
loaded_sheet = loaded_workbook.active

row_data = [loaded_sheet.cell(row=1, column=i).value for i in range(1, 11)]
print("Первая строка данных:", row_data)

column_data = [loaded_sheet.cell(row=2, column=i).value for i in range(1, 11)]
print("Вторая строка данных:", column_data)

new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

for i, value in enumerate(row_data, start=1):
    new_sheet.cell(row=i, column=1, value=value)

new_workbook.save('vertical_data.xlsx')